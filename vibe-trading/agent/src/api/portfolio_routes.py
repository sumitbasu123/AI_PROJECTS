"""Personal Indian-equities portfolio API with paper-only order execution."""

from __future__ import annotations

import json
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Literal

from fastapi import APIRouter, Depends, FastAPI, HTTPException
from pydantic import BaseModel, Field

PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_PORTFOLIO_FILE = PROJECT_ROOT / "portfolio_holdings.xlsx"
ORDER_LEDGER = Path(__file__).resolve().parents[2] / "data" / "paper_orders.json"


class PaperOrderRequest(BaseModel):
    symbol: str = Field(..., min_length=1, max_length=40)
    side: Literal["buy", "sell"]
    quantity: int = Field(..., gt=0, le=1_000_000)
    order_type: Literal["market", "limit"] = "market"
    limit_price: float | None = Field(None, gt=0)
    product: Literal["delivery", "intraday"] = "delivery"
    execution_mode: Literal["simulated", "connector_paper"] = "simulated"
    connector_profile: str | None = None
    note: str = Field("", max_length=500)


def _portfolio_path() -> Path:
    configured = os.getenv("VIBE_TRADING_PORTFOLIO_FILE", "").strip()
    return Path(configured).expanduser().resolve() if configured else DEFAULT_PORTFOLIO_FILE


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _load_holdings(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"portfolio workbook not found: {path}")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    headers = {str(value).strip().upper(): index for index, value in enumerate(rows[0]) if value is not None}
    required = {"INSTRUMENT", "NET POS", "AVG COST", "LTP"}
    missing = required - headers.keys()
    if missing:
        raise ValueError(f"portfolio workbook is missing columns: {', '.join(sorted(missing))}")

    def cell(row: tuple[Any, ...], name: str) -> Any:
        index = headers.get(name)
        return row[index] if index is not None and index < len(row) else None

    holdings: list[dict[str, Any]] = []
    for row in rows[1:]:
        symbol = str(cell(row, "INSTRUMENT") or "").strip().upper()
        if not symbol:
            continue
        quantity = _number(cell(row, "NET POS"))
        average_cost = _number(cell(row, "AVG COST"))
        ltp = _number(cell(row, "LTP"))
        invested = _number(cell(row, "INV AMT")) or quantity * average_cost
        market_value = _number(cell(row, "MKT VAL")) or quantity * ltp
        total_pnl = _number(cell(row, "TOTAL P&L")) or market_value - invested
        holdings.append(
            {
                "symbol": symbol,
                "quantity": quantity,
                "average_cost": average_cost,
                "ltp": ltp,
                "change_pct": _number(cell(row, "CHG %")),
                "invested_value": invested,
                "market_value": market_value,
                "total_pnl": total_pnl,
                "total_pnl_pct": total_pnl / invested if invested else 0.0,
                "day_pnl": _number(cell(row, "P&L DAY")),
            }
        )
    return holdings


def _portfolio_summary() -> dict[str, Any]:
    path = _portfolio_path()
    holdings = _load_holdings(path)
    invested = sum(row["invested_value"] for row in holdings)
    market_value = sum(row["market_value"] for row in holdings)
    total_pnl = sum(row["total_pnl"] for row in holdings)
    day_pnl = sum(row["day_pnl"] for row in holdings)
    for row in holdings:
        row["weight"] = row["market_value"] / market_value if market_value else 0.0
    holdings.sort(key=lambda item: item["market_value"], reverse=True)

    top_weight = holdings[0]["weight"] if holdings else 0.0
    hhi = sum(row["weight"] ** 2 for row in holdings)
    day_start_value = market_value - day_pnl
    return {
        "as_of": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(),
        "source_file": str(path),
        "currency": "INR",
        "metrics": {
            "positions": len(holdings),
            "invested_value": invested,
            "market_value": market_value,
            "total_pnl": total_pnl,
            "total_pnl_pct": total_pnl / invested if invested else 0.0,
            "day_pnl": day_pnl,
            "day_pnl_pct": day_pnl / day_start_value if day_start_value else 0.0,
            "top_position_weight": top_weight,
            "concentration_hhi": hhi,
        },
        "holdings": holdings,
        "risk_flags": _risk_flags(holdings, top_weight, hhi),
    }


def _risk_flags(holdings: list[dict[str, Any]], top_weight: float, hhi: float) -> list[dict[str, str]]:
    flags: list[dict[str, str]] = []
    if top_weight > 0.20 and holdings:
        flags.append(
            {
                "level": "warning",
                "title": "Single-position concentration",
                "detail": f"{holdings[0]['symbol']} is {top_weight:.1%} of current market value.",
            }
        )
    if hhi > 0.15:
        flags.append(
            {
                "level": "warning",
                "title": "Portfolio concentration",
                "detail": f"HHI is {hhi:.3f}; review sector and issuer overlap before adding exposure.",
            }
        )
    losers = [row for row in holdings if row["total_pnl_pct"] <= -0.20]
    if losers:
        flags.append(
            {
                "level": "info",
                "title": "Deep drawdown review",
                "detail": f"{len(losers)} holding(s) are at least 20% below average cost.",
            }
        )
    return flags


def _read_orders() -> list[dict[str, Any]]:
    if not ORDER_LEDGER.exists():
        return []
    try:
        payload = json.loads(ORDER_LEDGER.read_text(encoding="utf-8"))
        return payload if isinstance(payload, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def _write_orders(orders: list[dict[str, Any]]) -> None:
    ORDER_LEDGER.parent.mkdir(parents=True, exist_ok=True)
    temp = ORDER_LEDGER.with_suffix(".tmp")
    temp.write_text(json.dumps(orders, indent=2) + "\n", encoding="utf-8")
    temp.replace(ORDER_LEDGER)


def _india_connectors() -> list[dict[str, Any]]:
    from src.trading.profiles import list_profiles

    return [
        {
            "id": profile.id,
            "label": profile.label,
            "connector": profile.connector,
            "environment": profile.environment,
            "readonly": profile.readonly,
            "capabilities": list(profile.capabilities),
            "notes": profile.notes,
        }
        for profile in list_profiles()
        if profile.connector in {"dhan", "shoonya"}
    ]


def register_portfolio_routes(app: FastAPI, require_auth: Callable[..., Any]) -> None:
    router = APIRouter(prefix="/portfolio", tags=["portfolio"], dependencies=[Depends(require_auth)])

    @router.get("/summary")
    async def portfolio_summary() -> dict[str, Any]:
        try:
            return _portfolio_summary()
        except (FileNotFoundError, ValueError) as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

    @router.get("/connectors")
    async def portfolio_connectors() -> dict[str, Any]:
        return {"connectors": _india_connectors(), "live_ordering_enabled": False}

    @router.get("/orders")
    async def list_paper_orders() -> dict[str, Any]:
        return {"orders": list(reversed(_read_orders()))}

    @router.post("/orders")
    async def create_paper_order(payload: PaperOrderRequest) -> dict[str, Any]:
        if payload.order_type == "limit" and payload.limit_price is None:
            raise HTTPException(status_code=422, detail="limit_price is required for limit orders")
        if payload.execution_mode == "connector_paper" and not payload.connector_profile:
            raise HTTPException(status_code=422, detail="connector_profile is required for connector paper orders")

        symbol = payload.symbol.strip().upper()
        result: dict[str, Any] = {
            "status": "simulated",
            "order_status": "recorded",
            "paper_guard": "local_ledger",
        }
        if payload.execution_mode == "connector_paper":
            from src.trading.profiles import profile_by_id
            from src.trading.service import place_order

            profile = profile_by_id(payload.connector_profile)
            if profile.connector not in {"dhan", "shoonya"} or profile.environment != "paper":
                raise HTTPException(status_code=403, detail="only Dhan or Shoonya paper profiles are allowed")
            if "orders.place" not in profile.capabilities:
                raise HTTPException(status_code=403, detail="selected connector profile is read-only")
            result = place_order(
                symbol,
                profile_id=profile.id,
                side=payload.side,
                quantity=payload.quantity,
                order_type=payload.order_type,
                limit_price=payload.limit_price,
                time_in_force="day",
            )
            if result.get("status") != "ok":
                raise HTTPException(status_code=422, detail=result.get("error", "paper connector rejected order"))

        order = {
            "id": str(uuid.uuid4()),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "symbol": symbol,
            "side": payload.side,
            "quantity": payload.quantity,
            "order_type": payload.order_type,
            "limit_price": payload.limit_price,
            "product": payload.product,
            "execution_mode": payload.execution_mode,
            "connector_profile": payload.connector_profile,
            "note": payload.note.strip(),
            "result": result,
        }
        orders = _read_orders()
        orders.append(order)
        _write_orders(orders[-1000:])
        return order

    app.include_router(router)

